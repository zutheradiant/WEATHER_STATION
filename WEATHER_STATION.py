import tkinter as tk
from tkinter import messagebox
from PIL import Image, ImageTk
import adafruit_dht
import board
import threading
import time
from picamera2 import Picamera2
import os
from openpyxl import Workbook, load_workbook

# Define the folder path
data_folder = 'data/'

# Check if the folder exists, if not, create it
if not os.path.exists(data_folder):
    os.makedirs(data_folder)

# Set the Excel file path to the folder
excel_path = os.path.join(data_folder, 'weather_station_data.xlsx')

# Check if the Excel file exists, if not, create it
if not os.path.exists(excel_path):
    # Create a new workbook and add headers
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Weather Data'
    sheet.append(['Timestamp', 'Temperature (C)', 'Humidity (%)'])  # Add header row
    workbook.save(excel_path)

# Function to save data to Excel
def save_data_to_excel(temperature, humidity):
    workbook = load_workbook(excel_path)
    sheet = workbook.active
    timestamp = time.strftime('%Y-%m-%d %H:%M:%S')  # Current time
    sheet.append([timestamp, temperature, humidity])
    workbook.save(excel_path)

# Sensor setup
sensor = adafruit_dht.DHT11(board.D4)  # Using GPIO pin 4 for the DHT 11 sensor

# Camera setup
picam2 = Picamera2()
camera_config = picam2.create_still_configuration()
picam2.configure(camera_config)
picam2.start()

# Tkinter setup
root = tk.Tk()
root.title("Mini Weather Station")
root.attributes('-fullscreen', True)

# Labels for temperature and humidity
temp_label = tk.Label(root, text="Temperature: --.- C", font=("Helvetica", 16))
temp_label.pack(pady=10)

humidity_label = tk.Label(root, text="Humidity: --.- %", font=("Helvetica", 16))
humidity_label.pack(pady=10)


# Canvas for live camera feed
canvas = tk.Canvas(root, width=640, height=360)
canvas.pack()

# Function to update sensor data
def update_sensor_data():
    while True:
        try:
            temperature = sensor.temperature
            humidity = sensor.humidity
            if humidity is not None and temperature is not None:
                temp_label.config(text=f"Temperature: {temperature:.1f} C")
                humidity_label.config(text=f"Humidity: {humidity:.1f} %")
                # Save data to Excel
                save_data_to_excel(temperature, humidity)
            else:
                temp_label.config(text="Failed to get reading. Try again!")
                humidity_label.config(text="")
        except RuntimeError as error:
            print(f"Sensor reading error: {error}")
        time.sleep(10)  # Read every minute

# Function to update live camera feed
def update_camera_feed():
    while True:
        img_path = 'weather_image.jpg'
        picam2.capture_file(img_path)
        img = Image.open(img_path)
        img = img.resize((640, 360), Image.ANTIALIAS)
        img_tk = ImageTk.PhotoImage(img)
        canvas.create_image(0, 0, anchor=tk.NW, image=img_tk)
        canvas.image = img_tk
        time.sleep(0.1)

# Function to capture and save image
def snap_and_save():
    img_path = 'snap_image.jpg'
    picam2.capture_file(img_path)
    messagebox.showinfo("Image Captured", f"Image saved as {os.path.basename(img_path)}")

# Button to snap and save image
snap_button = tk.Button(root, text="Snap & Save Picture", command=snap_and_save)
snap_button.pack(pady=20)

# Start sensor and camera feed threads
sensor_thread = threading.Thread(target=update_sensor_data)
sensor_thread.daemon = True
sensor_thread.start()

camera_thread = threading.Thread(target=update_camera_feed)
camera_thread.daemon = True
camera_thread.start()

root.mainloop()
root.bind("<escape>", exit)
